"""
Generador de datos de prueba para IA Contable
Crea un archivo Excel con transacciones contables ficticias
"""
from datetime import datetime, timedelta
import random
from pathlib import Path

# Intenta usar openpyxl directamente si está disponible
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

def generar_datos_excel():
    """Genera datos en formato Excel"""
    if not EXCEL_AVAILABLE:
        return None
    
    clientes = ["Empresa A", "Empresa B", "Consultores XYZ", "Tech Solutions", "Distribuidora Central"]
    proveedores = ["Proveedor 1", "Proveedor 2", "Mayorista Nacional", "Importador Sur"]
    productos = ["Servicio Consultoría", "Producto A", "Producto B", "Licencias Software", "Soporte Técnico"]

    start_date = datetime.now() - timedelta(days=60)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Transacciones"
    
    # Headers
    headers = ["factura", "fecha", "cliente", "producto", "cantidad", "precio_unitario", "subtotal", "iva", "total", "tipo"]
    ws.append(headers)
    
    # Formatear header
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    num_ventas = 0
    num_compras = 0
    ventas_total = 0
    compras_total = 0
    
    for i in range(80):
        fecha = start_date + timedelta(days=random.randint(0, 60))
        
        if random.choice([True, False]):  # Venta
            factura_num = f"FAC-2025-{i:04d}"
            cliente = random.choice(clientes)
            monto = random.uniform(500000, 5000000)
            tipo = "Factura de Venta"
            num_ventas += 1
            ventas_total += monto
        else:  # Compra
            factura_num = f"COM-2025-{i:04d}"
            cliente = random.choice(proveedores)
            monto = random.uniform(300000, 2000000)
            tipo = "Factura de Compra"
            num_compras += 1
            compras_total += monto
        
        cantidad = random.randint(1, 10)
        precio_unit = monto / cantidad
        iva = monto * 0.19
        total = monto + iva
        
        ws.append([
            factura_num,
            fecha.strftime("%Y-%m-%d"),
            cliente,
            random.choice(productos),
            cantidad,
            f"{precio_unit:.2f}",
            f"{monto:.2f}",
            f"{iva:.2f}",
            f"{total:.2f}",
            tipo
        ])
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    # Guardar
    excel_path = Path("data/transacciones_contables.xlsx")
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_path)
    
    print(f"✅ Archivo generado: {excel_path}")
    print(f"📊 Total de transacciones: 80")
    print(f"💰 Monto total: ${(ventas_total + compras_total) * 1.19:,.0f} COP")
    print(f"📈 Ventas: {num_ventas} facturas = ${ventas_total:,.0f}")
    print(f"📦 Compras: {num_compras} facturas = ${compras_total:,.0f}")
    print(f"📅 Rango: Últimos 60 días")
    
    return excel_path

if __name__ == "__main__":
    print("� Generando datos de prueba para IA Contable...")
    print()
    
    # Generar Excel
    excel_path = generar_datos_excel()
    
    if excel_path:
        print()
        print("✅ Datos generados exitosamente")
        print("📂 Puedes subirlos a la app a través de Streamlit")
    else:
        print("❌ openpyxl no está disponible")
        print("Instala con: pip install openpyxl")
